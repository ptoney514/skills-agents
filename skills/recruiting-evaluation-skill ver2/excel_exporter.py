"""
Excel Export Module for Recruiting Evaluation Skill

Generates formatted Excel reports from SQLite database.

Usage:
    from recruiting_database import RecruitingDatabase
    from excel_exporter import export_candidate_scores, export_shortlist

    db = RecruitingDatabase('recruiting.db')
    export_candidate_scores(db, job_id=1, output_file='candidate_scores.xlsx')
"""

from typing import Optional, List, Dict, Any
from pathlib import Path
from datetime import datetime

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    print("Warning: openpyxl not available. Excel export will be disabled.")
    print("Install with: pip install openpyxl")


def export_candidate_scores(db, job_id: int, output_file: str = 'candidate_scores.xlsx'):
    """
    Export all candidate scores to Excel with formatting.

    Args:
        db: RecruitingDatabase instance
        job_id: Job ID to export
        output_file: Output file path
    """
    if not OPENPYXL_AVAILABLE:
        print("ERROR: openpyxl is required for Excel export")
        return

    wb = Workbook()

    # Sheet 1: Candidate Scores
    ws = wb.active
    ws.title = "Candidate Scores"

    # Get data
    candidates = db.get_candidates_with_scores(job_id)
    stats = db.get_summary_stats(job_id)
    job = db.get_active_job()

    # Headers
    headers = [
        "Rank", "Candidate Name", "Resume File", "Last Position",
        "Q Score", "E Score", "R Score", "Overall Score",
        "Stars", "Recommendation", "Current Stage", "Evaluated Date",
        "Strengths", "Concerns"
    ]

    ws.append(headers)

    # Format header row
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Add data rows
    for rank, candidate in enumerate(candidates, 1):
        # Format strengths and concerns
        strengths_text = ""
        if candidate.get('strengths'):
            strengths_text = "\n".join(f"• {s}" for s in candidate['strengths'])

        concerns_text = ""
        if candidate.get('concerns'):
            concerns_text = "\n".join(f"• {c}" for c in candidate['concerns'])

        # Format stars
        stars = "⭐" * (candidate.get('stars') or 0)

        # Format evaluation date
        eval_date = ""
        if candidate.get('evaluation_date'):
            try:
                dt = datetime.fromisoformat(candidate['evaluation_date'])
                eval_date = dt.strftime('%Y-%m-%d %I:%M %p')
            except:
                eval_date = candidate['evaluation_date']

        row_data = [
            rank,
            candidate.get('full_name', ''),
            candidate.get('resume_filename', ''),
            candidate.get('last_position', ''),
            candidate.get('q_score'),
            candidate.get('e_score'),
            candidate.get('r_score'),
            candidate.get('overall_score'),
            stars,
            candidate.get('recommendation', ''),
            candidate.get('current_stage', ''),
            eval_date,
            strengths_text,
            concerns_text
        ]

        ws.append(row_data)

        # Apply conditional formatting to overall score
        row_num = rank + 1
        score_cell = ws.cell(row=row_num, column=8)
        score = candidate.get('overall_score', 0)

        if score >= 90:
            score_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        elif score >= 85:
            score_cell.fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        elif score >= 70:
            score_cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        elif score > 0:
            score_cell.fill = PatternFill(start_color="F4CCCC", end_color="F4CCCC", fill_type="solid")

    # Auto-adjust column widths
    column_widths = {
        'A': 6,   # Rank
        'B': 25,  # Name
        'C': 30,  # Resume File
        'D': 35,  # Last Position
        'E': 10,  # Q Score
        'F': 10,  # E Score
        'G': 10,  # R Score
        'H': 12,  # Overall
        'I': 12,  # Stars
        'J': 20,  # Recommendation
        'K': 15,  # Stage
        'L': 20,  # Date
        'M': 50,  # Strengths
        'N': 50,  # Concerns
    }

    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    # Enable text wrapping for strengths and concerns
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=13, max_col=14):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical='top')

    # Freeze header row
    ws.freeze_panes = 'A2'

    # Enable autofilter
    ws.auto_filter.ref = ws.dimensions

    # Sheet 2: Summary Statistics
    ws_summary = wb.create_sheet("Summary")

    # Job info
    ws_summary.append(["Job Title", job.get('job_title', '') if job else ''])
    ws_summary.append(["Job Description", job.get('job_description_file', '') if job else ''])
    ws_summary.append(["Created Date", job.get('created_date', '') if job else ''])
    ws_summary.append([])

    # Statistics
    ws_summary.append(["Candidate Summary"])
    ws_summary.append(["Total Candidates", stats['total_candidates']])
    ws_summary.append(["Strong Candidates (90+)", stats['strong_candidates']])
    ws_summary.append(["Good Candidates (85-89)", stats['good_candidates']])
    ws_summary.append(["Potential Fit (70-84)", stats['potential_fit']])
    ws_summary.append(["Not a Match (<70)", stats['no_match']])
    ws_summary.append([])

    # Stage breakdown
    ws_summary.append(["Candidates by Stage"])
    for stage, count in stats['by_stage'].items():
        ws_summary.append([stage.replace('_', ' ').title(), count])

    ws_summary.append([])
    ws_summary.append(["Exported", datetime.now().strftime('%Y-%m-%d %I:%M %p')])

    # Format summary sheet
    for cell in ws_summary['A']:
        cell.font = Font(bold=True)

    ws_summary.column_dimensions['A'].width = 30
    ws_summary.column_dimensions['B'].width = 20

    # Save workbook
    wb.save(output_file)
    print(f"✅ Exported candidate scores to {output_file}")


def export_shortlist(db, job_id: int, stage: str, output_file: str):
    """
    Export shortlist for a specific stage.

    Args:
        db: RecruitingDatabase instance
        job_id: Job ID
        stage: Stage to export (phone_screen, interview, final_round)
        output_file: Output file path
    """
    if not OPENPYXL_AVAILABLE:
        print("ERROR: openpyxl is required for Excel export")
        return

    wb = Workbook()
    ws = wb.active
    ws.title = f"{stage.replace('_', ' ').title()} Candidates"

    # Get shortlist data
    candidates_in_stage = db.get_shortlist_by_stage(job_id, stage)

    # Get their evaluations
    candidate_data = []
    for candidate in candidates_in_stage:
        eval_data = db.get_latest_evaluation(candidate['candidate_id'])
        notes = db.get_notes(candidate['candidate_id'])

        candidate_data.append({
            **candidate,
            'evaluation': eval_data,
            'notes_count': len(notes)
        })

    # Headers
    headers = [
        "Rank", "Candidate Name", "Overall Score", "Stars",
        "Recommendation", "Added to Stage", "Decision Rationale",
        "Notes Count", "Last Position"
    ]

    ws.append(headers)

    # Format headers
    for cell in ws[1]:
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal='center')

    # Sort by overall score
    candidate_data.sort(key=lambda x: x['evaluation']['overall_score'] if x.get('evaluation') else 0, reverse=True)

    # Add data
    for rank, data in enumerate(candidate_data, 1):
        eval_data = data.get('evaluation', {})

        stars = "⭐" * (eval_data.get('stars', 0))

        added_date = ""
        if data.get('added_date'):
            try:
                dt = datetime.fromisoformat(data['added_date'])
                added_date = dt.strftime('%Y-%m-%d')
            except:
                added_date = data['added_date']

        ws.append([
            rank,
            data.get('full_name', ''),
            eval_data.get('overall_score', ''),
            stars,
            eval_data.get('recommendation', ''),
            added_date,
            data.get('decision_rationale', ''),
            data.get('notes_count', 0),
            eval_data.get('last_position', '')
        ])

    # Column widths
    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 40
    ws.column_dimensions['H'].width = 12
    ws.column_dimensions['I'].width = 35

    # Enable wrapping for rationale
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=7, max_col=7):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical='top')

    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = ws.dimensions

    wb.save(output_file)
    print(f"✅ Exported {stage} shortlist to {output_file}")


def export_comparison_report(db, job_id: int, candidate_ids: List[int], output_file: str):
    """
    Export side-by-side comparison of finalists.

    Args:
        db: RecruitingDatabase instance
        job_id: Job ID
        candidate_ids: List of candidate IDs to compare
        output_file: Output file path
    """
    if not OPENPYXL_AVAILABLE:
        print("ERROR: openpyxl is required for Excel export")
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "Finalist Comparison"

    # Get candidate data
    candidates_data = []
    for cid in candidate_ids:
        candidate = db.get_candidate(cid)
        evaluation = db.get_latest_evaluation(cid)
        notes = db.get_notes(cid)
        eval_history = db.get_evaluation_history(cid)

        candidates_data.append({
            'candidate': candidate,
            'evaluation': evaluation,
            'notes': notes,
            'eval_history': eval_history
        })

    # Build comparison table
    rows = [
        ["Metric"] + [data['candidate']['full_name'] for data in candidates_data],
        [],
        ["SCORES"],
        ["Overall Score"] + [data['evaluation']['overall_score'] for data in candidates_data],
        ["Stars"] + ["⭐" * data['evaluation']['stars'] for data in candidates_data],
        ["Recommendation"] + [data['evaluation']['recommendation'] for data in candidates_data],
        [],
        ["Qualifications (Q)"] + [data['evaluation']['q_score'] for data in candidates_data],
        ["Experience (E)"] + [data['evaluation']['e_score'] for data in candidates_data],
        ["Risk Flags (R)"] + [data['evaluation']['r_score'] for data in candidates_data],
        [],
        ["DETAILS"],
        ["Last Position"] + [data['evaluation']['last_position'] for data in candidates_data],
        ["Current Stage"] + [data['candidate']['current_stage'] for data in candidates_data],
        ["Notes Count"] + [len(data['notes']) for data in candidates_data],
        ["Evaluations"] + [len(data['eval_history']) for data in candidates_data],
    ]

    for row in rows:
        ws.append(row)

    # Formatting
    ws.column_dimensions['A'].width = 20
    for i in range(len(candidates_data)):
        col_letter = get_column_letter(i + 2)
        ws.column_dimensions[col_letter].width = 25

    # Bold first column
    for cell in ws['A']:
        cell.font = Font(bold=True)

    # Section headers
    for row in [3, 12]:
        for cell in ws[row]:
            cell.font = Font(bold=True, size=12)
            cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")

    # Add strengths and concerns
    current_row = ws.max_row + 2

    for idx, data in enumerate(candidates_data):
        col = idx + 2
        col_letter = get_column_letter(col)

        ws[f'{col_letter}{current_row}'] = "STRENGTHS"
        ws[f'{col_letter}{current_row}'].font = Font(bold=True)

        strengths = data['evaluation'].get('strengths', [])
        for i, strength in enumerate(strengths, 1):
            ws[f'{col_letter}{current_row + i}'] = f"• {strength}"

        concerns_row = current_row + len(strengths) + 2
        ws[f'{col_letter}{concerns_row}'] = "CONCERNS"
        ws[f'{col_letter}{concerns_row}'].font = Font(bold=True)

        concerns = data['evaluation'].get('concerns', [])
        for i, concern in enumerate(concerns, 1):
            ws[f'{col_letter}{concerns_row + i + 1}'] = f"• {concern}"

    wb.save(output_file)
    print(f"✅ Exported comparison report to {output_file}")


if __name__ == "__main__":
    print("Excel Exporter Module")
    print("This module provides Excel export functions for recruiting database.")
    print("\nUsage:")
    print("  from recruiting_database import RecruitingDatabase")
    print("  from excel_exporter import export_candidate_scores")
    print("  db = RecruitingDatabase('recruiting.db')")
    print("  export_candidate_scores(db, job_id=1, output_file='candidates.xlsx')")
